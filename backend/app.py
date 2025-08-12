from datetime import datetime, timedelta, date
from dateutil.relativedelta import relativedelta
from flask import Flask, jsonify, request, send_file
from flask_cors import CORS
import io
import locale
import math
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill, NamedStyle, numbers


app = Flask(__name__)
CORS(app)


DATE_FMT = "%d/%m/%Y"


def parse_date(value: str) -> date:
    return datetime.strptime(value, DATE_FMT).date()


def last_day_of_month(d: date) -> date:
    first_next_month = (d.replace(day=1) + relativedelta(months=+1))
    return first_next_month - timedelta(days=1)


def month_name_es(d: date) -> str:
    # Intentar usar locale español, con fallback
    try:
        locale.setlocale(locale.LC_TIME, "es_ES.UTF-8")
    except locale.Error:
        try:
            locale.setlocale(locale.LC_TIME, "es_ES")
        except locale.Error:
            # Fallback manual
            nombres = [
                "enero",
                "febrero",
                "marzo",
                "abril",
                "mayo",
                "junio",
                "julio",
                "agosto",
                "septiembre",
                "octubre",
                "noviembre",
                "diciembre",
            ]
            return nombres[d.month - 1].capitalize()
    return d.strftime("%B").capitalize()


def daterange_monthly(start: date, end: date):
    """Yield (del, hasta) pares mensuales de periodos cerrados, el último puede terminar en 'end'.

    - del: fecha inicio inclusive
    - hasta: fecha fin inclusive
    """
    current_start = start
    while current_start <= end:
        # Alinear al calendario: cerrar en el último día del mes del current_start
        month_end = last_day_of_month(current_start)
        current_end = min(month_end, end)
        yield current_start, current_end
        current_start = current_end + timedelta(days=1)


def days_in_month_of(d: date) -> int:
    return last_day_of_month(d).day


def calculate_rows(start: date, end: date, base: float, monthly_rate_pct: float):
    rows = []
    total_interest = 0
    for dt_start, dt_end in daterange_monthly(start, end):
        # Días efectivos del período
        days = (dt_end - dt_start).days + 1
        # Fórmula solicitada: proporcional a días sobre base 30
        # Interés = base * tasa_mensual * (días / 30)
        interest = base * (monthly_rate_pct / 100.0) * (days / 30.0)
        interest_rounded = int(round(interest))
        total_interest += interest_rounded
        rows.append(
            {
                "mes": month_name_es(dt_start),
                "del": dt_start.strftime(DATE_FMT),
                "hasta": dt_end.strftime(DATE_FMT),
                "dias": days,
                "base": base,
                "tasa": monthly_rate_pct,
                "interes": interest_rounded,
            }
        )
    return rows, int(total_interest)


def generate_tramos(start: date, end: date, base: float, monthly_rate_pct: float, vencimiento: date | None):
    if vencimiento is None or not (start < vencimiento <= end):
        rows, total = calculate_rows(start, end, base, monthly_rate_pct)
        return {
            "tramos": [
                {
                    "titulo": "TABLA DE LIQUIDACIÓN GENERAL DEL CRÉDITO",
                    "rows": rows,
                    "subtotal": total,
                }
            ],
            "total": total,
        }

    # Tramo 1: vigencia hasta el día anterior al vencimiento
    tramo1_end = vencimiento - timedelta(days=1)
    rows1, subtotal1 = calculate_rows(start, tramo1_end, base, monthly_rate_pct)
    # Calcular duración contractual en meses para el título dinámico
    rd = relativedelta(vencimiento, start)
    meses_credito = rd.years * 12 + rd.months
    # Si cae exactamente en el mismo día del mes (día 0), mantener; si hay desfase positivo, considerar redondeo hacia arriba
    if rd.days > 0:
        meses_credito += 1

    # Tramo 2: desde el vencimiento hasta el corte
    rows2, subtotal2 = calculate_rows(vencimiento, end, base, monthly_rate_pct)

    return {
        "tramos": [
            {
                "titulo": f"TABLA DE LIQUIDACIÓN GENERAL DEL CRÉDITO DE HIPOTECA {meses_credito} MESES",
                "rows": rows1,
                "subtotal": subtotal1,
            },
            {
                "titulo": "TABLA DE LIQUIDACIÓN GENERAL DEL CRÉDITO DE HIPOTECA DESDE QUE SE VENCE EL PLAZO PACTADO",
                "rows": rows2,
                "subtotal": subtotal2,
            },
        ],
        "total": int(subtotal1 + subtotal2),
    }


@app.route("/api/calculate", methods=["POST"])
def api_calculate():
    data = request.get_json(force=True)
    start = parse_date(data["fechaInicial"])  # dd/mm/yyyy
    end = parse_date(data["fechaCorte"])  # dd/mm/yyyy
    base = float(data["capitalBase"])  # number
    tasa = float(data["tasaMensual"])  # porcentaje mensual
    vencimiento = None
    if data.get("fechaVencimiento"):
        vencimiento = parse_date(data["fechaVencimiento"])

    result = generate_tramos(start, end, base, tasa, vencimiento)
    return jsonify(result)


def build_excel(payload: dict) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Liquidación"

    thin = Side(border_style="thin", color="000000")
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

    currency_style = NamedStyle(name="currency_style")
    currency_style.number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
    currency_style.border = border_all

    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    bold = Font(bold=True)

    # Comienzo en la fila 10 y columna 2 (B10)
    row_idx = 10
    start_col = 2
    for idx, tramo in enumerate(payload["tramos"], start=1):
        # Título tramo
        title = tramo["titulo"]
        ws.merge_cells(start_row=row_idx, start_column=start_col, end_row=row_idx, end_column=start_col + 6)
        cell = ws.cell(row=row_idx, column=start_col, value=title)
        cell.font = Font(bold=True, size=12)
        cell.alignment = Alignment(horizontal="center")
        row_idx += 1

        # Encabezados
        headers = [
            "Mes causado",
            "Del",
            "Hasta",
            "Días",
            "Base",
            "Tasa de interés moratorio",
            "Interés causado",
        ]
        for offset, h in enumerate(headers):
            c = ws.cell(row=row_idx, column=start_col + offset, value=h)
            c.font = bold
            c.fill = header_fill
            c.border = border_all
            c.alignment = Alignment(horizontal="center")
        row_idx += 1

        # Filas
        for r in tramo["rows"]:
            ws.cell(row=row_idx, column=start_col + 0, value=r["mes"]).border = border_all
            ws.cell(row=row_idx, column=start_col + 1, value=r["del"]).border = border_all
            ws.cell(row=row_idx, column=start_col + 2, value=r["hasta"]).border = border_all
            ws.cell(row=row_idx, column=start_col + 3, value=r["dias"]).border = border_all
            base_cell = ws.cell(row=row_idx, column=start_col + 4, value=float(r["base"]))
            base_cell.style = currency_style
            tasa_cell = ws.cell(row=row_idx, column=start_col + 5, value=f"{r['tasa']:.2f}%")
            tasa_cell.border = border_all
            tasa_cell.alignment = Alignment(horizontal="center")
            interes_cell = ws.cell(row=row_idx, column=start_col + 6, value=float(r["interes"]))
            interes_cell.style = currency_style
            row_idx += 1

        # Subtotal tramo
        ws.merge_cells(start_row=row_idx, start_column=start_col, end_row=row_idx, end_column=start_col + 5)
        st_cell = ws.cell(row=row_idx, column=start_col, value="Subtotal tramo")
        st_cell.font = bold
        st_cell.border = border_all
        st_val = ws.cell(row=row_idx, column=start_col + 6, value=float(tramo["subtotal"]))
        st_val.style = currency_style
        row_idx += 2

    # Total general
    ws.merge_cells(start_row=row_idx, start_column=start_col, end_row=row_idx, end_column=start_col + 5)
    total_cell = ws.cell(row=row_idx, column=start_col, value="Total intereses causados a esta fecha")
    total_cell.font = Font(bold=True, size=12)
    total_cell.border = border_all
    total_val = ws.cell(row=row_idx, column=start_col + 6, value=float(payload["total"]))
    total_val.style = currency_style

    # Ajustar anchos
    widths = [18, 12, 12, 8, 14, 26, 16]
    for i, w in enumerate(widths):
        col_idx = start_col + i
        ws.column_dimensions[chr(64 + col_idx)].width = w

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream


@app.route("/api/export", methods=["POST"])
def api_export():
    data = request.get_json(force=True)
    start = parse_date(data["fechaInicial"])  # dd/mm/yyyy
    end = parse_date(data["fechaCorte"])  # dd/mm/yyyy
    base = float(data["capitalBase"])  # number
    tasa = float(data["tasaMensual"])  # porcentaje mensual
    vencimiento = None
    if data.get("fechaVencimiento"):
        vencimiento = parse_date(data["fechaVencimiento"])

    payload = generate_tramos(start, end, base, tasa, vencimiento)
    stream = build_excel(payload)

    filename = f"liquidacion_{start.strftime('%Y%m%d')}_{end.strftime('%Y%m%d')}.xlsx"
    return send_file(
        stream,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)


