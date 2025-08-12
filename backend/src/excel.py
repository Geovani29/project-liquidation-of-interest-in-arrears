from __future__ import annotations

import io
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill, NamedStyle, numbers


def build_excel(payload: dict) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Liquidación"

    thin = Side(border_style="thin", color="000000")
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

    currency_style = NamedStyle(name="currency_cop")
    currency_style.number_format = '"$"#,##0'
    currency_style.border = border_all
    if "currency_cop" not in wb.named_styles:
        wb.add_named_style(currency_style)

    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    bold = Font(bold=True)

    row_idx = 10
    start_col = 2
    for idx, tramo in enumerate(payload["tramos"], start=1):
        title = tramo["titulo"]
        ws.merge_cells(start_row=row_idx, start_column=start_col, end_row=row_idx, end_column=start_col + 6)
        cell = ws.cell(row=row_idx, column=start_col, value=title)
        cell.font = Font(bold=True, size=12)
        cell.alignment = Alignment(horizontal="center")
        row_idx += 1

        headers = [
            "Mes causado",
            "Del",
            "Hasta",
            "Días",
            "Base",
            "Tasa de interés moratorio",
            "Interés causado",
        ]
        for offset, h in enumerate(headers):
            c = ws.cell(row=row_idx, column=start_col + offset, value=h)
            c.font = bold
            c.fill = header_fill
            c.border = border_all
            c.alignment = Alignment(horizontal="center")
        row_idx += 1

        for r in tramo["rows"]:
            ws.cell(row=row_idx, column=start_col + 0, value=r["mes"]).border = border_all
            ws.cell(row=row_idx, column=start_col + 1, value=r["del"]).border = border_all
            ws.cell(row=row_idx, column=start_col + 2, value=r["hasta"]).border = border_all
            ws.cell(row=row_idx, column=start_col + 3, value=r["dias"]).border = border_all
            base_cell = ws.cell(row=row_idx, column=start_col + 4, value=float(r["base"]))
            base_cell.style = currency_style
            tasa_cell = ws.cell(row=row_idx, column=start_col + 5, value=f"{r['tasa']:.2f}%")
            tasa_cell.border = border_all
            tasa_cell.alignment = Alignment(horizontal="center")
            interes_cell = ws.cell(row=row_idx, column=start_col + 6, value=float(r["interes"]))
            interes_cell.style = currency_style
            row_idx += 1

        ws.merge_cells(start_row=row_idx, start_column=start_col, end_row=row_idx, end_column=start_col + 5)
        st_cell = ws.cell(row=row_idx, column=start_col, value="Subtotal tramo")
        st_cell.font = bold
        st_cell.border = border_all
        st_val = ws.cell(row=row_idx, column=start_col + 6, value=float(tramo["subtotal"]))
        st_val.style = currency_style
        row_idx += 2

    ws.merge_cells(start_row=row_idx, start_column=start_col, end_row=row_idx, end_column=start_col + 5)
    total_cell = ws.cell(row=row_idx, column=start_col, value="Total intereses causados a esta fecha")
    total_cell.font = Font(bold=True, size=12)
    total_cell.border = border_all
    total_val = ws.cell(row=row_idx, column=start_col + 6, value=float(payload["total"]))
    total_val.style = currency_style

    widths = [18, 12, 12, 8, 14, 26, 16]
    for i, w in enumerate(widths):
        col_idx = start_col + i
        ws.column_dimensions[chr(64 + col_idx)].width = w

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream


